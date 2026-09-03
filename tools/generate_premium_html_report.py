"""Generate a premium, self-contained HTML monthly project health report.

Reads AI2C_Day2_Monthly_Report_Data_Pack_Demo.xlsx (or compatible 8-sheet pack)
and produces one HTML file with all images inlined as base64 — safe to email
to management with no missing-asset risk.

Usage:
    python tools/generate_premium_html_report.py
        --input "AI2C_Day2_Monthly_Report_Data_Pack_Demo.xlsx"
        --period 2026-04
        --output "outputs/monthly_report/2026-04/monthly_report_premium.html"
"""

from __future__ import annotations

import argparse
import base64
import html
import io
import os
import sys
from dataclasses import dataclass
from datetime import datetime, date
from pathlib import Path
from typing import Any, Iterable

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import openpyxl


# -----------------------------------------------------------------------------
# Brand & RAG palette
# -----------------------------------------------------------------------------

BRAND_ORANGE = "#F36B15"
BRAND_PURPLE = "#9B0A68"
BRAND_DARK = "#231F20"
BRAND_LIGHT = "#FAFAFA"

RAG_COLOR = {
    "Green": "#1F9D55",
    "Yellow": "#E5A823",
    "Red": "#D7263D",
    "Data Gap": "#6B7280",
    "Unknown": "#9CA3AF",
}

RAG_LIGHT = {
    "Green": "#E6F4EC",
    "Yellow": "#FBEDD9",
    "Red": "#FBE3E7",
    "Data Gap": "#EEF0F3",
    "Unknown": "#F1F2F4",
}

CRITICAL_PATH_HIGHLIGHT = "#7A0F47"

# Plot styling
plt.rcParams.update({
    "font.family": "DejaVu Sans",
    "font.size": 10,
    "axes.titleweight": "bold",
    "axes.titlesize": 12,
    "axes.edgecolor": "#cccccc",
    "axes.labelcolor": BRAND_DARK,
    "xtick.color": BRAND_DARK,
    "ytick.color": BRAND_DARK,
    "axes.spines.top": False,
    "axes.spines.right": False,
})


# -----------------------------------------------------------------------------
# Data loading
# -----------------------------------------------------------------------------

def load_workbook_as_dict(xlsx_path: Path) -> dict[str, list[dict[str, Any]]]:
    """Read an Excel workbook into {sheet_name: list-of-row-dicts}.

    Cell values are taken as-is (data_only=True, so cached formula values
    are used)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    out: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 1:
            out[sheet_name] = []
            continue
        headers = [c.value for c in ws[1]]
        rows: list[dict[str, Any]] = []
        for r in range(2, ws.max_row + 1):
            row = {}
            for i, h in enumerate(headers):
                if h is None:
                    continue
                row[str(h)] = ws.cell(r, i + 1).value
            # skip fully-empty rows
            if any(v not in (None, "") for v in row.values()):
                rows.append(row)
        out[sheet_name] = rows
    return out


def to_date_str(v: Any) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v)


def safe_float(v: Any) -> float | None:
    if v in (None, ""):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def pct(v: Any, digits: int = 1) -> str:
    f = safe_float(v)
    if f is None:
        return "—"
    return f"{f * 100:.{digits}f}%"


# -----------------------------------------------------------------------------
# KPI computation
# -----------------------------------------------------------------------------

@dataclass
class KPIs:
    progress_red: int
    progress_amber: int
    progress_green: int
    progress_data_gap: int
    progress_avg_variance: float | None

    milestone_red: int
    milestone_amber: int
    milestone_green: int
    milestone_critical_red: int
    milestone_max_delay_days: int

    rfi_open: int
    rfi_overdue: int
    rfi_red: int
    rfi_max_overdue: int

    procurement_red: int
    procurement_amber: int
    procurement_data_gap: int
    procurement_max_delay: int

    safety_red: int
    safety_open: int
    safety_overdue: int

    quality_open_ncr: int

    commercial_current_gp_pct: float | None
    commercial_objective_gp_pct: float | None
    commercial_gp_delta_hkd_m: float | None
    commercial_rag: str

    decision_red: int
    decision_amber: int
    decision_total_open: int

    overall_status: str  # GO / CONTROL / STOP


def compute_kpis(data: dict[str, list[dict[str, Any]]]) -> KPIs:
    progress = data.get("01_Area_Progress", [])
    milestones = data.get("02_Programme_Milestones", [])
    rfi = data.get("03_Submission_RFI", [])
    procurement = data.get("04_Procurement", [])
    safety = data.get("05_Safety_Quality", [])
    commercial = data.get("06_Commercial_Cost", [])
    decisions = data.get("07_Risk_Action_Decision", [])

    def count_rag(rows, level):
        return sum(1 for r in rows if str(r.get("RAG_Status", "")).strip() == level)

    def count_risk_status(rows, level):
        return sum(1 for r in rows if str(r.get("Risk_Status", "")).strip() == level)

    # Progress variance: average actual-vs-planned variance across rows that have both
    p_var = [safe_float(r.get("Variance_%")) for r in progress]
    p_var = [v for v in p_var if v is not None]
    progress_avg_variance = sum(p_var) / len(p_var) if p_var else None

    # Milestones
    m_red = count_rag(milestones, "Red")
    m_amber = count_rag(milestones, "Yellow")
    m_green = count_rag(milestones, "Green")
    m_crit_red = sum(
        1 for r in milestones
        if str(r.get("RAG_Status", "")).strip() == "Red"
        and str(r.get("Critical_Path", "")).strip().lower() == "yes"
    )
    m_delays = [safe_float(r.get("Variance_Days")) for r in milestones]
    m_delays = [int(d) for d in m_delays if d is not None]
    m_max_delay = max(m_delays) if m_delays else 0

    # RFI
    rfi_open = sum(
        1 for r in rfi
        if str(r.get("Status_Normalized", "")).strip() in {"Open", "Pending", "Not Yet Submitted"}
    )
    rfi_overdue = sum(
        1 for r in rfi
        if (safe_float(r.get("Days_Open_Overdue")) or 0) > 0
        and str(r.get("Status_Normalized", "")).strip() in {"Open", "Pending"}
    )
    rfi_red = count_rag(rfi, "Red")
    rfi_max_overdue_vals = [
        safe_float(r.get("Days_Open_Overdue")) or 0 for r in rfi
        if str(r.get("Status_Normalized", "")).strip() in {"Open", "Pending"}
    ]
    rfi_max_overdue = int(max(rfi_max_overdue_vals)) if rfi_max_overdue_vals else 0

    # Procurement
    proc_red = count_risk_status(procurement, "Red")
    proc_amber = count_risk_status(procurement, "Yellow")
    proc_gap = count_risk_status(procurement, "Data Gap")
    proc_var = [safe_float(r.get("Variance_Days")) for r in procurement]
    proc_var = [int(v) for v in proc_var if v is not None]
    proc_max_delay = max(proc_var) if proc_var else 0

    # Safety
    safety_red = count_rag(safety, "Red")
    safety_open = sum(1 for r in safety if str(r.get("Status", "")).strip() == "Open")
    safety_overdue = sum(
        1 for r in safety
        if (safe_float(r.get("Days_Overdue")) or 0) > 0
    )
    quality_open_ncr = sum(
        1 for r in safety
        if str(r.get("Category", "")).strip() == "Quality"
        and str(r.get("Status", "")).strip() == "Open"
    )

    # Commercial - latest row
    current_row = commercial[-1] if commercial else {}
    c_current = safe_float(current_row.get("Current_GP_%"))
    c_obj = safe_float(current_row.get("Obj_GP_%"))
    c_delta = safe_float(current_row.get("GP_vs_Objective_HKD_M"))
    c_rag = str(current_row.get("RAG_Status", "")).strip() or "Unknown"

    # Decisions
    d_red = sum(
        1 for r in decisions
        if str(r.get("RAG_Status", "")).strip() == "Red"
        and str(r.get("Status", "")).strip() == "Open"
    )
    d_amber = sum(
        1 for r in decisions
        if str(r.get("RAG_Status", "")).strip() == "Yellow"
        and str(r.get("Status", "")).strip() == "Open"
    )
    d_open = sum(1 for r in decisions if str(r.get("Status", "")).strip() == "Open")

    # Overall status
    if safety_red > 0 or m_crit_red > 0 or proc_red > 0:
        overall = "STOP"
    elif rfi_red > 0 or d_red > 0 or count_rag(progress, "Red") > 0 or c_rag == "Red":
        overall = "STOP" if d_red >= 3 else "CONTROL"
    elif c_rag == "Yellow" or m_amber + count_rag(progress, "Yellow") > 0:
        overall = "CONTROL"
    else:
        overall = "GO"

    return KPIs(
        progress_red=count_rag(progress, "Red"),
        progress_amber=count_rag(progress, "Yellow"),
        progress_green=count_rag(progress, "Green"),
        progress_data_gap=count_rag(progress, "Data Gap"),
        progress_avg_variance=progress_avg_variance,
        milestone_red=m_red,
        milestone_amber=m_amber,
        milestone_green=m_green,
        milestone_critical_red=m_crit_red,
        milestone_max_delay_days=m_max_delay,
        rfi_open=rfi_open,
        rfi_overdue=rfi_overdue,
        rfi_red=rfi_red,
        rfi_max_overdue=rfi_max_overdue,
        procurement_red=proc_red,
        procurement_amber=proc_amber,
        procurement_data_gap=proc_gap,
        procurement_max_delay=proc_max_delay,
        safety_red=safety_red,
        safety_open=safety_open,
        safety_overdue=safety_overdue,
        quality_open_ncr=quality_open_ncr,
        commercial_current_gp_pct=c_current,
        commercial_objective_gp_pct=c_obj,
        commercial_gp_delta_hkd_m=c_delta,
        commercial_rag=c_rag,
        decision_red=d_red,
        decision_amber=d_amber,
        decision_total_open=d_open,
        overall_status=overall,
    )


# -----------------------------------------------------------------------------
# AI summary (rule-based, deterministic)
# -----------------------------------------------------------------------------

def build_ai_summary(data, kpis: KPIs, period: str) -> list[str]:
    """Produce 5-8 management-language bullet findings.

    Deterministic — same data in, same bullets out. This is the 'AI Drafts'
    voice but implemented as auditable rules, not a free-text LLM call."""
    findings: list[str] = []

    # Headline: programme + critical path
    if kpis.milestone_critical_red > 0:
        findings.append(
            f"Critical-path risk: {kpis.milestone_critical_red} milestone(s) on the "
            f"critical path are Red, with the worst delay at {kpis.milestone_max_delay_days} days. "
            "Recovery plan must be on the Day-1 agenda."
        )
    elif kpis.milestone_red > 0:
        findings.append(
            f"Programme: {kpis.milestone_red} milestone(s) Red but none on critical path. "
            f"Maximum slippage observed is {kpis.milestone_max_delay_days} days."
        )

    # Progress
    if kpis.progress_red > 0:
        findings.append(
            f"Progress: {kpis.progress_red} Red activity row(s) require management attention; "
            f"average variance to plan is {pct(kpis.progress_avg_variance, 1)}."
        )

    # RFI / Submission
    if kpis.rfi_red > 0 or kpis.rfi_max_overdue >= 5:
        findings.append(
            f"Submission / RFI: {kpis.rfi_overdue} open item(s) overdue, worst case "
            f"{kpis.rfi_max_overdue} day(s) past response-due. Technical Manager escalation needed."
        )

    # Procurement
    if kpis.procurement_red > 0 or kpis.procurement_max_delay >= 7:
        findings.append(
            f"Procurement: {kpis.procurement_red} Red item(s), with longest delivery slip "
            f"of {kpis.procurement_max_delay} day(s) versus required-on-site date. Supplier "
            "recovery meeting required."
        )

    # Safety
    if kpis.safety_red > 0:
        findings.append(
            f"Safety: {kpis.safety_red} Red item(s) outstanding. "
            "Closure evidence must be provided before report issue."
        )

    # Commercial
    if kpis.commercial_rag in {"Yellow", "Red"}:
        findings.append(
            f"Commercial: current GP at {pct(kpis.commercial_current_gp_pct, 2)} versus "
            f"objective {pct(kpis.commercial_objective_gp_pct, 2)} "
            f"(delta HK${kpis.commercial_gp_delta_hkd_m:+.2f}M). "
            "QS to review VO pricing and commitment exposure."
        )

    # Decisions
    if kpis.decision_red > 0:
        findings.append(
            f"Decisions: {kpis.decision_red} Red decision item(s) open in the action log — "
            "Director sign-off required this cycle."
        )

    # Data Quality
    if kpis.progress_data_gap + kpis.procurement_data_gap > 0:
        findings.append(
            f"Data quality: {kpis.progress_data_gap + kpis.procurement_data_gap} record(s) "
            "have data gaps. AI did NOT infer missing values — values left blank for human update."
        )

    # Status badge
    findings.append(
        f"Overall management status: **{kpis.overall_status}**. "
        "AI drafted this report; human review gates must be cleared before external issue."
    )

    return findings


# -----------------------------------------------------------------------------
# Charts → base64 PNG
# -----------------------------------------------------------------------------

def fig_to_base64(fig) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode("ascii")


def chart_rag_donut(data) -> str:
    """Total RAG mix across all categories."""
    sheets = [
        "01_Area_Progress",
        "02_Programme_Milestones",
        "03_Submission_RFI",
        "05_Safety_Quality",
        "07_Risk_Action_Decision",
    ]
    counts = {"Green": 0, "Yellow": 0, "Red": 0, "Data Gap": 0}
    for s in sheets:
        for row in data.get(s, []):
            v = str(row.get("RAG_Status", "")).strip()
            if v in counts:
                counts[v] += 1
    # Add Procurement (different field name)
    for row in data.get("04_Procurement", []):
        v = str(row.get("Risk_Status", "")).strip()
        if v in counts:
            counts[v] += 1

    labels = [k for k, v in counts.items() if v > 0]
    sizes = [counts[k] for k in labels]
    colors = [RAG_COLOR[k] for k in labels]
    total = sum(sizes) or 1

    fig, ax = plt.subplots(figsize=(6.4, 5.2))
    wedges, _ = ax.pie(
        sizes,
        colors=colors,
        startangle=90,
        wedgeprops=dict(width=0.38, edgecolor="white", linewidth=2),
    )

    # Centre label: total count
    ax.text(0, 0.06, f"{total}", ha="center", va="center",
            fontsize=30, fontweight="bold", color=BRAND_DARK)
    ax.text(0, -0.14, "items", ha="center", va="center",
            fontsize=10, color="#666")

    # Per-wedge labels OUTSIDE the donut with leader lines, so nothing
    # gets clipped by the donut hole and every count is fully readable.
    for wedge, label, size in zip(wedges, labels, sizes):
        ang = (wedge.theta2 + wedge.theta1) / 2.0
        x = 1.18 * np.cos(np.deg2rad(ang))
        y = 1.18 * np.sin(np.deg2rad(ang))
        ha = "left" if x > 0 else "right"
        # leader line
        ax.annotate(
            f"{label}\n{size} ({size / total * 100:.0f}%)",
            xy=(0.96 * np.cos(np.deg2rad(ang)), 0.96 * np.sin(np.deg2rad(ang))),
            xytext=(x, y),
            ha=ha, va="center",
            fontsize=11, color=BRAND_DARK,
            arrowprops=dict(arrowstyle="-", color="#999", lw=0.8,
                            connectionstyle="arc3,rad=0.0"),
            bbox=dict(boxstyle="round,pad=0.3", fc=RAG_LIGHT.get(label, "white"),
                      ec="none"),
        )

    ax.set_xlim(-1.8, 1.8)
    ax.set_ylim(-1.5, 1.5)
    ax.set_aspect("equal")
    ax.set_title("RAG Mix Across All Project Dimensions",
                 pad=14, color=BRAND_PURPLE, fontsize=13)
    return fig_to_base64(fig)


def chart_progress_planned_vs_actual(data) -> str:
    rows = data.get("01_Area_Progress", [])
    labels = []
    planned = []
    actual = []
    rags = []
    for r in rows:
        p = safe_float(r.get("Planned_%_This_Month"))
        a = safe_float(r.get("Actual_%_This_Month"))
        labels.append(f"{r.get('Progress_ID')} {r.get('Zone_Area')}/{r.get('Trade')}")
        planned.append((p or 0) * 100)
        actual.append((a or 0) * 100)
        rags.append(str(r.get("RAG_Status", "")).strip() or "Unknown")

    fig, ax = plt.subplots(figsize=(10, max(4.8, 0.42 * len(rows) + 1.2)))
    y = list(range(len(labels)))
    ax.barh([i + 0.18 for i in y], planned, height=0.36,
            color="#D9D9D9", label="Planned %")
    bars = ax.barh([i - 0.18 for i in y], actual, height=0.36,
                   color=[RAG_COLOR.get(rg, RAG_COLOR["Unknown"]) for rg in rags],
                   label="Actual %")
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("% complete (this month cumulative)")
    ax.set_xlim(0, 110)
    ax.set_title("Area Progress — Planned vs Actual %, coloured by RAG", color=BRAND_PURPLE)
    ax.grid(axis="x", linestyle=":", alpha=0.5)

    legend_handles = [
        mpatches.Patch(color="#D9D9D9", label="Planned %"),
        mpatches.Patch(color=RAG_COLOR["Green"], label="Actual — Green"),
        mpatches.Patch(color=RAG_COLOR["Yellow"], label="Actual — Yellow"),
        mpatches.Patch(color=RAG_COLOR["Red"], label="Actual — Red"),
        mpatches.Patch(color=RAG_COLOR["Data Gap"], label="Actual — Data Gap"),
    ]
    ax.legend(handles=legend_handles, loc="lower right", frameon=False, fontsize=8)
    return fig_to_base64(fig)


def chart_milestone_variance(data) -> str:
    rows = data.get("02_Programme_Milestones", [])
    labels = []
    variances = []
    colors = []
    edge_colors = []
    for r in rows:
        v = safe_float(r.get("Variance_Days"))
        if v is None:
            continue
        labels.append(f"{r.get('Milestone_ID')} {r.get('Milestone_Name')[:30]}")
        variances.append(int(v))
        rag = str(r.get("RAG_Status", "")).strip() or "Unknown"
        colors.append(RAG_COLOR.get(rag, RAG_COLOR["Unknown"]))
        cp = str(r.get("Critical_Path", "")).strip().lower() == "yes"
        edge_colors.append(CRITICAL_PATH_HIGHLIGHT if cp else "white")

    fig, ax = plt.subplots(figsize=(10, max(4.5, 0.4 * len(rows) + 1)))
    y = list(range(len(labels)))
    bars = ax.barh(y, variances, color=colors, edgecolor=edge_colors, linewidth=2)
    ax.axvline(0, color="#555", linewidth=0.8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("Variance from Baseline (days) — negative = ahead, positive = behind")
    ax.set_title("Programme Milestones — Variance Days (thick purple edge = critical path)",
                 color=BRAND_PURPLE)
    ax.grid(axis="x", linestyle=":", alpha=0.5)

    legend_handles = [
        mpatches.Patch(color=RAG_COLOR["Green"], label="Green"),
        mpatches.Patch(color=RAG_COLOR["Yellow"], label="Yellow"),
        mpatches.Patch(color=RAG_COLOR["Red"], label="Red"),
        mpatches.Patch(facecolor="white", edgecolor=CRITICAL_PATH_HIGHLIGHT,
                       linewidth=2, label="Critical Path"),
    ]
    ax.legend(handles=legend_handles, loc="lower right", frameon=False, fontsize=8)
    return fig_to_base64(fig)


def chart_rfi_aging(data) -> str:
    rows = data.get("03_Submission_RFI", [])
    labels = []
    days = []
    colors = []
    for r in rows:
        d = safe_float(r.get("Days_Open_Overdue"))
        if d is None:
            d = 0
        labels.append(f"{r.get('Record_ID')} {r.get('Type','')[:14]}")
        days.append(int(d))
        rag = str(r.get("RAG_Status", "")).strip() or "Unknown"
        colors.append(RAG_COLOR.get(rag, RAG_COLOR["Unknown"]))

    fig, ax = plt.subplots(figsize=(10, max(4.0, 0.4 * len(rows) + 1)))
    y = list(range(len(labels)))
    ax.barh(y, days, color=colors)
    ax.axvline(0, color="#555", linewidth=0.8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("Days open or overdue (negative = ahead of due date)")
    ax.set_title("Submission / RFI Aging — Days Past Response-Due, coloured by RAG",
                 color=BRAND_PURPLE)
    ax.grid(axis="x", linestyle=":", alpha=0.5)
    return fig_to_base64(fig)


def chart_procurement_variance(data) -> str:
    rows = data.get("04_Procurement", [])
    labels = []
    variances = []
    colors = []
    for r in rows:
        v = safe_float(r.get("Variance_Days"))
        labels.append(f"{r.get('Item_ID')} {r.get('Item_Description','')[:24]}")
        variances.append(int(v) if v is not None else 0)
        rag = str(r.get("Risk_Status", "")).strip() or "Unknown"
        colors.append(RAG_COLOR.get(rag, RAG_COLOR["Unknown"]))

    fig, ax = plt.subplots(figsize=(10, max(4.0, 0.45 * len(rows) + 1)))
    y = list(range(len(labels)))
    ax.barh(y, variances, color=colors)
    ax.axvline(0, color="#555", linewidth=0.8)
    ax.set_yticks(y)
    ax.set_yticklabels(labels, fontsize=9)
    ax.invert_yaxis()
    ax.set_xlabel("Forecast delivery vs required on-site date (days)")
    ax.set_title("Procurement Variance — coloured by Risk Status", color=BRAND_PURPLE)
    ax.grid(axis="x", linestyle=":", alpha=0.5)
    return fig_to_base64(fig)


def chart_safety_quality(data) -> str:
    rows = data.get("05_Safety_Quality", [])
    categories = sorted({str(r.get("Category", "")).strip() for r in rows
                         if r.get("Category")})
    rag_levels = ["Green", "Yellow", "Red", "Data Gap"]
    counts = {cat: {lvl: 0 for lvl in rag_levels} for cat in categories}
    for r in rows:
        cat = str(r.get("Category", "")).strip()
        lvl = str(r.get("RAG_Status", "")).strip() or "Unknown"
        if cat in counts and lvl in counts[cat]:
            counts[cat][lvl] += 1

    fig, ax = plt.subplots(figsize=(8, 4.6))
    bottoms = [0] * len(categories)
    for lvl in rag_levels:
        vals = [counts[cat][lvl] for cat in categories]
        ax.bar(categories, vals, bottom=bottoms, color=RAG_COLOR[lvl], label=lvl,
               edgecolor="white", linewidth=1)
        bottoms = [bottoms[i] + vals[i] for i in range(len(categories))]
    ax.set_ylabel("Number of records")
    ax.set_title("Safety / Quality / Environmental — Records by RAG", color=BRAND_PURPLE)
    ax.legend(loc="upper right", frameon=False, fontsize=9)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    return fig_to_base64(fig)


def chart_commercial_gp(data) -> str:
    rows = data.get("06_Commercial_Cost", [])
    if not rows:
        fig, ax = plt.subplots(figsize=(7, 4))
        ax.text(0.5, 0.5, "No commercial data", ha="center")
        return fig_to_base64(fig)

    months = []
    obj_pct = []
    cur_pct = []
    obj_hkd = []
    cur_hkd = []
    for r in rows:
        d = r.get("Reporting_Month")
        months.append(d.strftime("%Y-%m") if isinstance(d, (datetime, date)) else str(d))
        obj_pct.append((safe_float(r.get("Obj_GP_%")) or 0) * 100)
        cur_pct.append((safe_float(r.get("Current_GP_%")) or 0) * 100)
        obj_hkd.append(safe_float(r.get("Obj_GP_HKD_M")) or 0)
        cur_hkd.append(safe_float(r.get("Current_GP_HKD_M")) or 0)

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(11, 4.4))
    x = list(range(len(months)))
    width = 0.36
    ax1.bar([i - width / 2 for i in x], obj_pct, width, color="#D9D9D9", label="Objective GP %")
    ax1.bar([i + width / 2 for i in x], cur_pct, width, color=BRAND_ORANGE, label="Current GP %")
    ax1.set_xticks(x)
    ax1.set_xticklabels(months)
    ax1.set_ylabel("GP %")
    ax1.set_title("GP % — Current vs Objective", color=BRAND_PURPLE)
    ax1.legend(frameon=False, fontsize=9)
    ax1.grid(axis="y", linestyle=":", alpha=0.5)

    ax2.bar([i - width / 2 for i in x], obj_hkd, width, color="#D9D9D9", label="Objective GP (HK$M)")
    ax2.bar([i + width / 2 for i in x], cur_hkd, width, color=BRAND_PURPLE, label="Current GP (HK$M)")
    ax2.set_xticks(x)
    ax2.set_xticklabels(months)
    ax2.set_ylabel("HK$ Million")
    ax2.set_title("GP HK$ — Current vs Objective", color=BRAND_PURPLE)
    ax2.legend(frameon=False, fontsize=9)
    ax2.grid(axis="y", linestyle=":", alpha=0.5)

    fig.tight_layout()
    return fig_to_base64(fig)


def chart_cashflow(data) -> str:
    rows = data.get("06_Commercial_Cost", [])
    if not rows:
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.text(0.5, 0.5, "No commercial data", ha="center")
        return fig_to_base64(fig)

    current = rows[-1]
    labels = [
        "Original\nContract", "Contract\nincl. VO",
        "Cost\nCertified", "Commitment",
        "Payment\nReceived", "Actual\nExpenditure",
    ]
    values = [
        safe_float(current.get("Original_Net_Contract_Sum_HKD_M")) or 0,
        safe_float(current.get("Contract_Sum_with_VO_HKD_M")) or 0,
        safe_float(current.get("Cost_Certified_HKD_M")) or 0,
        safe_float(current.get("Commitment_HKD_M")) or 0,
        safe_float(current.get("Payment_Received_HKD_M")) or 0,
        safe_float(current.get("Actual_Expenditure_HKD_M")) or 0,
    ]
    colors = [BRAND_DARK, BRAND_PURPLE, BRAND_ORANGE, "#C04A8B", "#1F9D55", "#E5A823"]

    fig, ax = plt.subplots(figsize=(10, 4.6))
    bars = ax.bar(labels, values, color=colors, edgecolor="white", linewidth=1.5)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width() / 2,
                bar.get_height() + 0.8,
                f"{val:.1f}", ha="center", fontsize=9, color=BRAND_DARK,
                fontweight="bold")
    ax.set_ylabel("HK$ Million")
    ax.set_title(f"Cashflow Snapshot — {to_date_str(current.get('Reporting_Month'))}",
                 color=BRAND_PURPLE)
    ax.set_ylim(0, max(values) * 1.18)
    ax.grid(axis="y", linestyle=":", alpha=0.5)
    return fig_to_base64(fig)


# -----------------------------------------------------------------------------
# HTML rendering
# -----------------------------------------------------------------------------

def file_to_base64(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode("ascii")


def rag_pill(value: str) -> str:
    v = str(value or "").strip()
    color = RAG_COLOR.get(v, "#9CA3AF")
    text_color = "white"
    return (
        f'<span class="pill" style="background:{color};color:{text_color}">'
        f'{html.escape(v or "—")}</span>'
    )


def status_pill(value: str) -> str:
    v = str(value or "").strip()
    color_map = {
        "Open": "#D7263D",
        "Pending": "#E5A823",
        "Pending approval": "#E5A823",
        "Pending Approval": "#E5A823",
        "Approved": "#1F9D55",
        "Answered": "#1F9D55",
        "Closed": "#1F9D55",
        "Not Yet Submitted": "#6B7280",
        "Issued": "#1F9D55",
        "Not Issued": "#D7263D",
        "Draft only": "#E5A823",
        "Pending Review": "#E5A823",
        "Blocked": "#D7263D",
    }
    c = color_map.get(v, "#6B7280")
    return f'<span class="pill" style="background:{c};color:white">{html.escape(v or "—")}</span>'


def render_table(
    rows: list[dict],
    columns: list[tuple[str, str, str]],
) -> str:
    """columns: list of (header, key, formatter_type).
    formatter_type one of: 'text','date','pct','rag','status','int','float2'."""
    out = ['<table class="data"><thead><tr>']
    for header, _, _ in columns:
        out.append(f"<th>{html.escape(header)}</th>")
    out.append("</tr></thead><tbody>")
    for row in rows:
        rag_value = str(row.get("RAG_Status") or row.get("Risk_Status") or "").strip()
        row_bg = RAG_LIGHT.get(rag_value, "")
        style = f' style="background:{row_bg}"' if row_bg else ""
        out.append(f"<tr{style}>")
        for _, key, ftype in columns:
            v = row.get(key)
            if ftype == "date":
                cell = html.escape(to_date_str(v))
            elif ftype == "pct":
                cell = pct(v, 1)
            elif ftype == "rag":
                cell = rag_pill(v)
            elif ftype == "status":
                cell = status_pill(v)
            elif ftype == "int":
                f = safe_float(v)
                cell = "—" if f is None else f"{int(f):,}"
            elif ftype == "float2":
                f = safe_float(v)
                cell = "—" if f is None else f"{f:,.2f}"
            else:
                cell = html.escape(str(v) if v not in (None, "") else "—")
            out.append(f"<td>{cell}</td>")
        out.append("</tr>")
    out.append("</tbody></table>")
    return "".join(out)


def kpi_card(title: str, big: str, sub: str, rag: str, icon: str = "") -> str:
    color = RAG_COLOR.get(rag, "#6B7280")
    return (
        '<div class="kpi-card" style="border-top:6px solid '
        f'{color}">'
        f'<div class="kpi-icon">{icon}</div>'
        f'<div class="kpi-title">{html.escape(title)}</div>'
        f'<div class="kpi-big">{html.escape(big)}</div>'
        f'<div class="kpi-sub">{html.escape(sub)}</div>'
        '</div>'
    )


def status_badge(overall: str) -> str:
    colors = {"GO": "#1F9D55", "CONTROL": "#E5A823", "STOP": "#D7263D"}
    label = {"GO": "GO — Confirm & Proceed",
             "CONTROL": "CONTROL — PM / QS / Planner Review",
             "STOP": "STOP — Director / Safety Escalation"}
    c = colors.get(overall, "#6B7280")
    return (
        f'<div class="status-badge" style="background:{c}">'
        f'{html.escape(label.get(overall, overall))}'
        '</div>'
    )


def section(title: str, body_html: str, anchor: str = "") -> str:
    anchor_attr = f' id="{anchor}"' if anchor else ""
    return (
        f'<section{anchor_attr}>'
        f'<h2>{html.escape(title)}</h2>'
        f'{body_html}'
        '</section>'
    )


CSS = """
:root {
  --brand-orange: #F36B15;
  --brand-purple: #9B0A68;
  --brand-dark:   #231F20;
  --brand-light:  #FAFAFA;
  --green:        #1F9D55;
  --yellow:       #E5A823;
  --red:          #D7263D;
  --gap:          #6B7280;
  --border:       #E5E7EB;
}
* { box-sizing: border-box; }
body {
  margin: 0;
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
  color: var(--brand-dark);
  background: #F4F5F7;
  -webkit-font-smoothing: antialiased;
}
.brand-strip {
  height: 12px;
  background: linear-gradient(90deg, var(--brand-orange) 0%, var(--brand-purple) 100%);
}
main {
  max-width: 1280px;
  margin: 0 auto;
  padding: 0 24px 80px;
}
.page-header {
  background: white;
  padding: 28px 36px;
  border-radius: 0 0 14px 14px;
  box-shadow: 0 2px 6px rgba(0,0,0,0.04);
  display: flex;
  justify-content: space-between;
  align-items: center;
  flex-wrap: wrap;
  gap: 16px;
}
.page-header .titles h1 {
  margin: 0 0 6px;
  font-size: 28px;
  color: var(--brand-dark);
  letter-spacing: -0.3px;
}
.page-header .titles .sub {
  color: #555;
  font-size: 14px;
}
.page-header .titles .sub strong { color: var(--brand-purple); }
.logo {
  max-height: 56px;
  width: auto;
}
.status-badge {
  color: white;
  font-weight: 700;
  font-size: 14px;
  padding: 10px 20px;
  border-radius: 999px;
  letter-spacing: 0.4px;
  box-shadow: 0 1px 4px rgba(0,0,0,0.12);
}

.meta-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
  gap: 14px;
  margin: 22px 0;
}
.meta-card {
  background: white;
  padding: 14px 18px;
  border-radius: 10px;
  box-shadow: 0 1px 3px rgba(0,0,0,0.04);
}
.meta-card .label {
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.6px;
  color: #888;
}
.meta-card .value {
  margin-top: 4px;
  font-size: 15px;
  font-weight: 600;
  color: var(--brand-dark);
}

.kpi-strip {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
  gap: 14px;
  margin-bottom: 28px;
}
.kpi-card {
  background: white;
  padding: 18px 18px 16px;
  border-radius: 12px;
  box-shadow: 0 2px 6px rgba(0,0,0,0.05);
  position: relative;
  overflow: hidden;
}
.kpi-card .kpi-icon {
  position: absolute;
  top: 16px;
  right: 16px;
  font-size: 22px;
  opacity: 0.18;
}
.kpi-title {
  font-size: 12px;
  text-transform: uppercase;
  letter-spacing: 0.5px;
  color: #888;
}
.kpi-big {
  margin-top: 6px;
  font-size: 30px;
  font-weight: 700;
  color: var(--brand-dark);
}
.kpi-sub {
  margin-top: 4px;
  font-size: 12px;
  color: #555;
  line-height: 1.4;
}

section {
  background: white;
  padding: 22px 28px;
  margin-bottom: 22px;
  border-radius: 12px;
  box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}
section h2 {
  margin: 0 0 16px;
  padding-bottom: 10px;
  border-bottom: 3px solid var(--brand-orange);
  color: var(--brand-purple);
  font-size: 18px;
  letter-spacing: -0.2px;
}
section h3 {
  color: var(--brand-dark);
  font-size: 15px;
  margin: 20px 0 8px;
}

.ai-summary {
  background: linear-gradient(135deg, #FFF7F0 0%, #FBEDF6 100%);
  border-left: 5px solid var(--brand-orange);
}
.ai-summary h2 {
  border-bottom-color: var(--brand-purple);
}
.ai-summary .ai-meta {
  font-size: 11px;
  text-transform: uppercase;
  letter-spacing: 0.7px;
  color: var(--brand-purple);
  font-weight: 700;
  margin-bottom: 10px;
}
.ai-summary ul {
  margin: 0;
  padding-left: 20px;
}
.ai-summary li {
  margin: 8px 0;
  line-height: 1.55;
  font-size: 14px;
  color: var(--brand-dark);
}
.ai-summary li strong { color: var(--brand-purple); }
.ai-summary .ai-footer {
  margin-top: 14px;
  font-size: 11px;
  color: #888;
  font-style: italic;
}

.chart {
  width: 100%;
  height: auto;
  border-radius: 8px;
  margin: 8px 0 4px;
}
.chart-caption {
  font-size: 12px;
  color: #666;
  margin-bottom: 14px;
}

table.data {
  width: 100%;
  border-collapse: collapse;
  font-size: 12.5px;
  margin: 6px 0 14px;
  border-radius: 8px;
  overflow: hidden;
}
table.data th {
  background: var(--brand-purple);
  color: white;
  text-align: left;
  padding: 9px 10px;
  font-weight: 600;
  font-size: 11.5px;
  text-transform: uppercase;
  letter-spacing: 0.4px;
}
table.data td {
  padding: 8px 10px;
  border-bottom: 1px solid #E5E7EB;
  vertical-align: top;
}
.pill {
  display: inline-block;
  padding: 2px 9px;
  font-size: 11px;
  font-weight: 600;
  border-radius: 999px;
  letter-spacing: 0.3px;
}

.gates-grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(260px, 1fr));
  gap: 12px;
}
.gate {
  border: 1px solid var(--border);
  border-radius: 10px;
  padding: 12px 14px;
  background: #fff;
}
.gate .gate-id {
  font-size: 11px;
  color: #888;
  text-transform: uppercase;
  letter-spacing: 0.5px;
}
.gate .gate-section {
  font-weight: 600;
  margin: 2px 0 6px;
  font-size: 14px;
}
.gate .gate-reviewer {
  font-size: 12px;
  color: #555;
  margin-bottom: 6px;
}

.decision-pack li {
  padding: 12px 14px;
  background: var(--brand-light);
  border-left: 4px solid var(--red);
  border-radius: 4px;
  margin: 8px 0;
  list-style: none;
  line-height: 1.5;
}
.decision-pack ul { padding: 0; }
.decision-pack li strong { color: var(--red); }

.footer {
  margin-top: 30px;
  text-align: center;
  font-size: 12px;
  color: #888;
  line-height: 1.6;
}
.footer .tagline {
  color: var(--brand-purple);
  font-weight: 600;
}

@media print {
  body { background: white; }
  section { box-shadow: none; border: 1px solid var(--border); page-break-inside: avoid; }
  .kpi-card, .meta-card { box-shadow: none; border: 1px solid var(--border); }
}
"""


def render_html(
    data: dict[str, list[dict[str, Any]]],
    kpis: KPIs,
    summary: list[str],
    period: str,
    source_path: Path,
    logo_b64: str | None,
    charts: dict[str, str],
) -> str:
    profile = {r["Field"]: r["Value"]
               for r in data.get("00_Project_Profile", [])
               if isinstance(r, dict) and r.get("Field")}

    project_name = profile.get("Project_Name", "—")
    project_id = profile.get("Project_ID", "—")
    contract_no = profile.get("Contract_No", "—")
    subcontractor = profile.get("Subcontractor", "—") or profile.get("Subcontractor_Name", "—")
    main_contractor = (profile.get("Main_Contractor") or profile.get("Client_Main_Contractor")
                       or profile.get("Client") or "—")
    report_as_of = profile.get("Report_As_Of", profile.get("Reporting_Date", period))

    # KPI strip
    cards = [
        kpi_card(
            "Programme Health",
            f"{kpis.milestone_red} Red",
            f"{kpis.milestone_critical_red} on critical path; max delay "
            f"{kpis.milestone_max_delay_days}d",
            "Red" if kpis.milestone_critical_red > 0 else
            ("Yellow" if kpis.milestone_amber > 0 else "Green"),
            "📅",
        ),
        kpi_card(
            "Area Progress",
            f"{kpis.progress_red} Red",
            f"{kpis.progress_amber} Yellow, {kpis.progress_data_gap} Data Gap; "
            f"avg variance {pct(kpis.progress_avg_variance, 1)}",
            "Red" if kpis.progress_red > 0 else
            ("Yellow" if kpis.progress_amber > 0 else "Green"),
            "📊",
        ),
        kpi_card(
            "Submission / RFI",
            f"{kpis.rfi_overdue} Overdue",
            f"{kpis.rfi_red} Red; worst {kpis.rfi_max_overdue} day(s) past due",
            "Red" if kpis.rfi_red > 0 else ("Yellow" if kpis.rfi_overdue > 0 else "Green"),
            "📨",
        ),
        kpi_card(
            "Procurement",
            f"{kpis.procurement_red} Red",
            f"{kpis.procurement_amber} Yellow; worst slip "
            f"{kpis.procurement_max_delay}d",
            "Red" if kpis.procurement_red > 0 else
            ("Yellow" if kpis.procurement_amber > 0 else "Green"),
            "🚚",
        ),
        kpi_card(
            "Safety / Quality",
            f"{kpis.safety_red} Red",
            f"{kpis.safety_open} open; {kpis.safety_overdue} overdue; "
            f"{kpis.quality_open_ncr} open NCR",
            "Red" if kpis.safety_red > 0 else
            ("Yellow" if kpis.safety_open > 0 else "Green"),
            "🦺",
        ),
        kpi_card(
            "Commercial",
            (f"{pct(kpis.commercial_current_gp_pct, 2)} GP"
             if kpis.commercial_current_gp_pct is not None else "—"),
            (f"Δ HK${kpis.commercial_gp_delta_hkd_m:+.2f}M vs objective"
             if kpis.commercial_gp_delta_hkd_m is not None else "no data"),
            kpis.commercial_rag if kpis.commercial_rag in RAG_COLOR else "Yellow",
            "💰",
        ),
    ]

    # AI Summary
    summary_li = []
    for s in summary:
        text = (s.replace("**", "")  # strip md bold tokens
                .replace("__", ""))
        # apply some inline bolding for emphasis
        rendered = html.escape(text)
        for token in ["GO", "CONTROL", "STOP"]:
            rendered = rendered.replace(token, f"<strong>{token}</strong>")
        summary_li.append(f"<li>{rendered}</li>")

    ai_summary_html = (
        '<div class="ai-meta">🤖 AI-Drafted Executive Summary · Period '
        f'{html.escape(period)}</div>'
        f'<ul>{"".join(summary_li)}</ul>'
        '<div class="ai-footer">Findings generated by deterministic rules over the '
        f'data pack. Source values were not paraphrased by the AI; numbers come '
        f'from the recomputed KPIs in {html.escape(source_path.name)}. '
        'Human review gates below must be cleared before external issue.</div>'
    )

    # Decision Pack (top Red items)
    decisions = data.get("07_Risk_Action_Decision", [])
    red_decisions = [d for d in decisions
                     if str(d.get("RAG_Status", "")).strip() == "Red"
                     and str(d.get("Status", "")).strip() == "Open"]
    decision_items = []
    for d in red_decisions:
        decision_items.append(
            f'<li><strong>{html.escape(str(d.get("Issue_ID", "")))}</strong> · '
            f'{html.escape(str(d.get("Category", "")))} — '
            f'{html.escape(str(d.get("Description", "")))} · '
            f'<em>Decision: {html.escape(str(d.get("Decision_Required", "")))}</em> · '
            f'Owner: {html.escape(str(d.get("Owner", "")))} · '
            f'Due: {to_date_str(d.get("Due_Date"))}</li>'
        )
    decision_pack_html = (
        '<div class="decision-pack">'
        '<p style="color:#555;margin:0 0 10px;">Red items requiring Director / '
        'specialist sign-off this cycle:</p>'
        f'<ul>{"".join(decision_items) or "<li>No Red decision items this cycle.</li>"}'
        '</ul></div>'
    )

    # Charts blocks
    def chart_block(title: str, b64: str, caption: str = "") -> str:
        return (
            f'<h3>{html.escape(title)}</h3>'
            f'<img class="chart" src="data:image/png;base64,{b64}" alt="{html.escape(title)}">'
            + (f'<div class="chart-caption">{html.escape(caption)}</div>' if caption else "")
        )

    # Detail tables
    progress_table = render_table(
        data.get("01_Area_Progress", []),
        [
            ("ID", "Progress_ID", "text"),
            ("Zone / Level / Trade", "Zone_Area", "text"),
            ("Activity", "Activity", "text"),
            ("Planned %", "Planned_%_This_Month", "pct"),
            ("Actual %", "Actual_%_This_Month", "pct"),
            ("Variance", "Variance_%", "pct"),
            ("RAG", "RAG_Status", "rag"),
            ("Reason", "Delay_Variance_Reason", "text"),
            ("Owner", "Owner", "text"),
            ("Due", "Due_Date", "date"),
        ],
    )

    milestone_table = render_table(
        data.get("02_Programme_Milestones", []),
        [
            ("ID", "Milestone_ID", "text"),
            ("Milestone", "Milestone_Name", "text"),
            ("Baseline", "Baseline_Date", "date"),
            ("Forecast", "Current_Forecast_Date", "date"),
            ("Δ days", "Variance_Days", "int"),
            ("Critical Path", "Critical_Path", "text"),
            ("RAG", "RAG_Status", "rag"),
            ("Reason", "Reason", "text"),
            ("Owner", "Owner", "text"),
        ],
    )

    rfi_table = render_table(
        data.get("03_Submission_RFI", []),
        [
            ("ID", "Record_ID", "text"),
            ("Type", "Type", "text"),
            ("Trade", "Package_Trade", "text"),
            ("Description", "Description", "text"),
            ("Days Open / Overdue", "Days_Open_Overdue", "int"),
            ("Status", "Status_Normalized", "status"),
            ("Programme Impact (days)", "Programme_Impact_Days", "int"),
            ("RAG", "RAG_Status", "rag"),
            ("Owner", "Responsible_Person", "text"),
        ],
    )

    procurement_table = render_table(
        data.get("04_Procurement", []),
        [
            ("ID", "Item_ID", "text"),
            ("Item", "Item_Description", "text"),
            ("Package", "Package", "text"),
            ("Required On-site", "Required_Onsite_Date", "date"),
            ("Approval", "Approval_Status", "status"),
            ("PO", "PO_Status", "status"),
            ("Forecast Delivery", "Forecast_Delivery_Date", "date"),
            ("Δ days", "Variance_Days", "int"),
            ("Risk", "Risk_Status", "rag"),
            ("Owner", "Owner", "text"),
        ],
    )

    safety_table = render_table(
        data.get("05_Safety_Quality", []),
        [
            ("ID", "Record_ID", "text"),
            ("Category", "Category", "text"),
            ("Metric / Issue", "Metric_or_Issue", "text"),
            ("Actual", "Actual_Result", "text"),
            ("Target", "Target", "text"),
            ("Status", "Status", "status"),
            ("Area", "Area", "text"),
            ("Days Overdue", "Days_Overdue", "int"),
            ("RAG", "RAG_Status", "rag"),
            ("Owner", "Owner", "text"),
        ],
    )

    decision_table = render_table(
        data.get("07_Risk_Action_Decision", []),
        [
            ("ID", "Issue_ID", "text"),
            ("Category", "Category", "text"),
            ("Linked Records", "Linked_Records", "text"),
            ("Description", "Description", "text"),
            ("Impact", "Impact", "text"),
            ("Decision Required", "Decision_Required", "text"),
            ("Owner", "Owner", "text"),
            ("Due", "Due_Date", "date"),
            ("Status", "Status", "status"),
            ("RAG", "RAG_Status", "rag"),
        ],
    )

    # Commercial cost: render as definition list because it's one wide row
    com_rows = data.get("06_Commercial_Cost", [])
    com = com_rows[-1] if com_rows else {}
    commercial_kv = [
        ("Original Contract (HK$M)", safe_float(com.get("Original_Net_Contract_Sum_HKD_M")), 2),
        ("Contract incl. VO (HK$M)", safe_float(com.get("Contract_Sum_with_VO_HKD_M")), 2),
        ("VO Net (HK$M)", safe_float(com.get("VO_Net_HKD_M")), 2),
        ("Objective GP %", safe_float(com.get("Obj_GP_%")), "pct"),
        ("Current GP %", safe_float(com.get("Current_GP_%")), "pct"),
        ("Δ GP HK$M", safe_float(com.get("GP_vs_Objective_HKD_M")), 2),
        ("Payment Received (HK$M)", safe_float(com.get("Payment_Received_HKD_M")), 2),
        ("Actual Expenditure (HK$M)", safe_float(com.get("Actual_Expenditure_HKD_M")), 2),
        ("Nett Cashflow (HK$M)", safe_float(com.get("Nett_Cashflow_HKD_M")), 2),
        ("Cost Certified (HK$M)", safe_float(com.get("Cost_Certified_HKD_M")), 2),
        ("Cost Certified %", safe_float(com.get("Cost_Certified_%")), "pct"),
        ("Commitment (HK$M)", safe_float(com.get("Commitment_HKD_M")), 2),
        ("Commitment %", safe_float(com.get("Commitment_%")), "pct"),
        ("Risk Exposure (HK$M)", safe_float(com.get("Risk_Exposure_HKD_M")), 2),
        ("Performance Score / 10", safe_float(com.get("Performance_Score_0_to_10")), 1),
    ]
    com_cards = []
    for label, v, fmt in commercial_kv:
        if v is None:
            display = "—"
        elif fmt == "pct":
            display = f"{v * 100:.2f}%"
        elif isinstance(fmt, int):
            display = f"{v:,.{fmt}f}"
        else:
            display = str(v)
        com_cards.append(
            f'<div class="meta-card"><div class="label">{html.escape(label)}</div>'
            f'<div class="value">{html.escape(display)}</div></div>'
        )
    commercial_grid = (
        '<div class="meta-grid">' + "".join(com_cards) + "</div>"
    )

    # Review gates
    review_rows = data.get("09_Review_Gates", [])
    gate_cards = []
    for g in review_rows:
        gate_cards.append(
            '<div class="gate">'
            f'<div class="gate-id">{html.escape(str(g.get("Review_Gate_ID","")))}</div>'
            f'<div class="gate-section">{html.escape(str(g.get("Report_Section","")))}</div>'
            f'<div class="gate-reviewer">Reviewer: <strong>'
            f'{html.escape(str(g.get("Reviewer","")))}</strong></div>'
            f'<div class="gate-reviewer">{html.escape(str(g.get("What_To_Check",""))[:160])}</div>'
            f'<div>{status_pill(g.get("Output_Status",""))}</div>'
            '</div>'
        )
    gates_html = f'<div class="gates-grid">{"".join(gate_cards)}</div>'

    # Logo
    logo_html = ""
    if logo_b64:
        logo_html = (
            f'<img class="logo" src="data:image/png;base64,{logo_b64}" '
            f'alt="OpenDeedigital">'
        )

    # Meta cards
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    meta_cards = [
        ('Reporting Period', period),
        ('Report As Of', to_date_str(report_as_of)),
        ('Contract No.', str(contract_no or "—")),
        ('Subcontractor', str(subcontractor or "—")),
        ('Main Contractor / Client', str(main_contractor or "—")),
        ('Generated', generated_at),
    ]
    meta_cards_html = "".join(
        f'<div class="meta-card"><div class="label">{html.escape(label)}</div>'
        f'<div class="value">{html.escape(value or "—")}</div></div>'
        for label, value in meta_cards
    )

    # Compose
    html_out = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{html.escape(project_name)} · Monthly Project Health Report · {html.escape(period)}</title>
  <style>{CSS}</style>
</head>
<body>
<div class="brand-strip"></div>
<main>

<div class="page-header">
  <div class="titles">
    <h1>{html.escape(project_name)}</h1>
    <div class="sub">
      <strong>Monthly Project Health Report</strong> ·
      Project <strong>{html.escape(project_id)}</strong> ·
      Period <strong>{html.escape(period)}</strong>
    </div>
  </div>
  <div style="display:flex; gap:18px; align-items:center;">
    {status_badge(kpis.overall_status)}
    {logo_html}
  </div>
</div>

<div class="meta-grid">
  {meta_cards_html}
</div>

<section class="ai-summary">
  <h2>AI Executive Summary · 月度執行摘要</h2>
  {ai_summary_html}
</section>

<h2 style="border:none;padding:0;margin:8px 4px 14px;color:#555;font-size:13px;
text-transform:uppercase;letter-spacing:0.6px;">Key Performance Indicators · 關鍵績效指標</h2>
<div class="kpi-strip">
  {"".join(cards)}
</div>

{section("Top Decision Pack · 須決策事項", decision_pack_html, "decisions")}

{section(
  "Project Health Dashboard · 項目健康儀表板",
  chart_block(
    "Overall RAG Mix Across All Dimensions",
    charts["rag"],
    "Combined RAG counts across Progress, Programme, RFI, Procurement, Safety, and Decisions."
  )
  + chart_block(
    "Area Progress — Planned vs Actual %",
    charts["progress"],
    "Bars coloured by RAG. Data Gap rows shown in grey — AI did not infer missing values."
  )
  + chart_block(
    "Programme Milestones — Variance Days",
    charts["milestones"],
    "Negative = ahead of baseline; positive = behind. Thick purple edge marks critical-path milestones."
  )
  + chart_block(
    "Submission / RFI Aging",
    charts["rfi"],
    "Days past response-due. Open + Pending items aged from raised date."
  )
  + chart_block(
    "Procurement Variance",
    charts["procurement"],
    "Forecast delivery date vs required on-site date, in days."
  )
  + chart_block(
    "Safety / Quality / Environmental — RAG Mix",
    charts["safety"],
    "Stacked counts of safety, quality, and environmental records by RAG."
  )
  + chart_block(
    "Commercial — GP % and GP HK$ (Current vs Objective)",
    charts["commercial"],
    "Side-by-side comparison across the latest reporting months."
  )
  + chart_block(
    "Cashflow Snapshot",
    charts["cashflow"],
    "Latest-period cashflow position across contract, certification, commitment, payment, and expenditure."
  ),
  "dashboard"
)}

{section("Area Progress · 區域施工進度", progress_table, "progress")}
{section("Programme Milestones · 程序里程碑", milestone_table, "programme")}
{section("Submission / RFI Register · 提交與 RFI 登記", rfi_table, "rfi")}
{section("Procurement Tracker · 採購追蹤", procurement_table, "procurement")}
{section("Safety / Quality / Environmental · 安全、品質、環境", safety_table, "safety")}
{section("Commercial Cost Summary · 商業成本摘要", commercial_grid, "commercial")}
{section("Risk · Action · Decision Log · 風險、行動、決策紀錄", decision_table, "actions")}

{section("Human Review Gates · 人手審閱閘門", gates_html, "gates")}

<div class="footer">
  <p class="tagline">AI Drafts · Human Checks · Human Signs</p>
  <p>This report was drafted automatically from <code>{html.escape(source_path.name)}</code>
  on {html.escape(generated_at)}. All KPI values were recomputed by deterministic Python tools;
  none were authored by a free-text language model. Human review gates above must be cleared
  before external issue.</p>
  <p>© OpenDeedigital × HKIC 2026 · AI2C Monthly Report Agent · Premium HTML edition</p>
</div>

</main>
</body>
</html>
"""
    return html_out


# -----------------------------------------------------------------------------
# Entry point
# -----------------------------------------------------------------------------

def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", default="demo-data/AI2C_Day2_Monthly_Report_Data_Pack_Demo.xlsx")
    parser.add_argument("--period", default="2026-04")
    parser.add_argument("--output", default=None)
    parser.add_argument("--logo", default="brand_assets/ODD Logo.png")
    args = parser.parse_args(argv)

    project_root = Path(__file__).resolve().parent.parent
    input_path = (project_root / args.input).resolve()
    logo_path = (project_root / args.logo).resolve()
    if args.output:
        out_path = (project_root / args.output).resolve()
    else:
        out_path = (
            project_root / "outputs" / "monthly_report" / args.period
            / "monthly_report_premium.html"
        ).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if not input_path.exists():
        print(f"ERROR: input not found: {input_path}", file=sys.stderr)
        return 1

    print(f"Loading workbook: {input_path}")
    data = load_workbook_as_dict(input_path)
    print(f"Loaded {len(data)} sheets: {list(data.keys())}")

    print("Computing KPIs ...")
    kpis = compute_kpis(data)
    print(f"  Overall: {kpis.overall_status}")
    print(f"  Progress R/A/G/Gap: {kpis.progress_red}/{kpis.progress_amber}/{kpis.progress_green}/{kpis.progress_data_gap}")
    print(f"  Milestones critical-Red: {kpis.milestone_critical_red}, max delay {kpis.milestone_max_delay_days}d")
    print(f"  RFI overdue: {kpis.rfi_overdue}, max overdue {kpis.rfi_max_overdue}d")
    print(f"  Procurement R: {kpis.procurement_red}, max delay {kpis.procurement_max_delay}d")
    print(f"  Safety R: {kpis.safety_red}, open {kpis.safety_open}")
    print(f"  Commercial: GP {pct(kpis.commercial_current_gp_pct, 2)} vs obj {pct(kpis.commercial_objective_gp_pct, 2)} (delta HK${kpis.commercial_gp_delta_hkd_m:+.2f}M)")

    print("Building AI summary ...")
    summary = build_ai_summary(data, kpis, args.period)
    print(f"  {len(summary)} bullets")

    print("Generating charts ...")
    charts = {
        "rag":         chart_rag_donut(data),
        "progress":    chart_progress_planned_vs_actual(data),
        "milestones":  chart_milestone_variance(data),
        "rfi":         chart_rfi_aging(data),
        "procurement": chart_procurement_variance(data),
        "safety":      chart_safety_quality(data),
        "commercial":  chart_commercial_gp(data),
        "cashflow":    chart_cashflow(data),
    }
    print(f"  {len(charts)} charts ready")

    logo_b64 = None
    if logo_path.exists():
        logo_b64 = file_to_base64(logo_path)
        print(f"  Logo embedded ({logo_path.stat().st_size:,} bytes)")
    else:
        print(f"  No logo at {logo_path}")

    print(f"Rendering HTML -> {out_path}")
    rendered = render_html(
        data=data,
        kpis=kpis,
        summary=summary,
        period=args.period,
        source_path=input_path,
        logo_b64=logo_b64,
        charts=charts,
    )
    out_path.write_text(rendered, encoding="utf-8")
    size_kb = out_path.stat().st_size / 1024
    print(f"  Wrote {size_kb:,.1f} KB")
    print()
    print(f"OK - open: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
