from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Dict, Iterable

import pandas as pd
from openpyxl import load_workbook

try:
    from ._common import project_path
except ImportError:  # pragma: no cover - supports direct script execution
    from _common import project_path


SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def normalize_name(value: object) -> str:
    """Normalize a workbook sheet, CSV stem, or column name for lookup."""
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def resolve_input_path(path: str | Path) -> Path:
    candidate = Path(path)
    if not candidate.is_absolute():
        candidate = project_path(str(candidate))
    return candidate.resolve()


def _unique_headers(raw_headers: Iterable[object]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(raw_headers, start=1):
        header = str(value).strip() if value is not None and str(value).strip() else f"Column_{index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        headers.append(header if count == 0 else f"{header}_{count + 1}")
    return headers


def _is_blank(value: object) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _worksheet_to_dataframe(ws) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    headers = _unique_headers(rows[0])
    data_rows = [list(row) for row in rows[1:] if any(not _is_blank(cell) for cell in row)]
    frame = pd.DataFrame(data_rows, columns=headers)

    keep_columns: list[str] = []
    for column in frame.columns:
        if not column.startswith("Column_") or frame[column].notna().any():
            keep_columns.append(column)
    return frame[keep_columns].reset_index(drop=True)


@dataclass(frozen=True)
class DataPack:
    source_path: Path
    sheets: Dict[str, pd.DataFrame]
    formula_sheets: Dict[str, pd.DataFrame]

    @property
    def sheet_names(self) -> list[str]:
        return list(self.sheets.keys())

    @property
    def normalized_sheet_names(self) -> dict[str, str]:
        return {normalize_name(name): name for name in self.sheets}

    def has_sheet(self, name: str) -> bool:
        return self.resolve_sheet_name(name) is not None

    def resolve_sheet_name(self, name: str) -> str | None:
        if name in self.sheets:
            return name
        return self.normalized_sheet_names.get(normalize_name(name))

    def get(self, name: str) -> pd.DataFrame:
        resolved = self.resolve_sheet_name(name)
        if resolved is None:
            raise KeyError(f"Sheet not found: {name}")
        return self.sheets[resolved].copy()

    def get_formulas(self, name: str) -> pd.DataFrame:
        resolved = self.resolve_sheet_name(name)
        if resolved is None:
            raise KeyError(f"Sheet not found: {name}")
        return self.formula_sheets.get(resolved, self.sheets[resolved]).copy()


def _load_excel(path: Path) -> DataPack:
    value_wb = load_workbook(path, data_only=True, read_only=False)
    formula_wb = load_workbook(path, data_only=False, read_only=False)
    value_sheets = {ws.title: _worksheet_to_dataframe(ws) for ws in value_wb.worksheets}
    formula_sheets = {ws.title: _worksheet_to_dataframe(ws) for ws in formula_wb.worksheets}
    return DataPack(source_path=path, sheets=value_sheets, formula_sheets=formula_sheets)


def _load_csv_folder(path: Path) -> DataPack:
    sheets: dict[str, pd.DataFrame] = {}
    for csv_path in sorted(path.glob("*.csv")):
        sheets[csv_path.stem] = pd.read_csv(csv_path)
    if not sheets:
        raise FileNotFoundError(f"No CSV files found in {path}")
    return DataPack(source_path=path, sheets=sheets, formula_sheets={k: v.copy() for k, v in sheets.items()})


def load_data_pack(path: str | Path) -> DataPack:
    resolved = resolve_input_path(path)
    if not resolved.exists():
        raise FileNotFoundError(f"Input data pack not found: {resolved}")

    if resolved.is_dir():
        return _load_csv_folder(resolved)
    if resolved.suffix.lower() == ".csv":
        frame = pd.read_csv(resolved)
        return DataPack(source_path=resolved, sheets={resolved.stem: frame}, formula_sheets={resolved.stem: frame.copy()})
    if resolved.suffix.lower() in SUPPORTED_EXCEL_SUFFIXES:
        return _load_excel(resolved)

    raise ValueError(f"Unsupported input type: {resolved.suffix or resolved}")


def get_profile_map(data_pack: DataPack) -> dict[str, object]:
    if not data_pack.has_sheet("00_Project_Profile"):
        return {}
    profile = data_pack.get("00_Project_Profile")
    if "Field" not in profile.columns or "Value" not in profile.columns:
        return {}
    return {
        str(row["Field"]).strip(): row["Value"]
        for _, row in profile.iterrows()
        if not pd.isna(row.get("Field")) and str(row.get("Field")).strip()
    }
